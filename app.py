import asyncio
import hashlib
import hmac
import io
import os
import uuid
import logging
import urllib.request
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Optional

import pandas as pd
import chardet
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.base import BaseHTTPMiddleware
from dotenv import load_dotenv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).parent / ".env", override=True)

import time

from normalizer.cache import CacheNormalizer
from normalizer.catalog import CatalogValidator
from normalizer.claude_client import ClaudeNormalizer
from normalizer.rules import clean_description, normalize_ref_candidate
from normalizer.db import init_db, Database

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DATA_DIR = Path(os.getenv("DATA_DIR", str(BASE_DIR / "data")))
API_KEY  = os.getenv("ANTHROPIC_API_KEY", "")
DB_HOST  = os.getenv("DB_HOST", "")
DB_USER  = os.getenv("DB_USER", "")
DB_PASS  = os.getenv("DB_PASS", "")
DB_NAME  = os.getenv("DB_NAME", "")
DB_PORT  = int(os.getenv("DB_PORT", "3306"))
MAX_BATCH_ROWS = int(os.getenv("MAX_BATCH_ROWS", "2000"))
COST_ALERT_EUR = float(os.getenv("COST_ALERT_EUR", "5.0"))

# Auth (opcional — si APP_PASSWORD y USERS están vacíos, la app queda abierta)
APP_PASSWORD = os.getenv("APP_PASSWORD", "")
SECRET_KEY   = os.getenv("SECRET_KEY", "skf-secret-default-changeme")
USERS_RAW    = os.getenv("USERS", "")   # formato: "user1:pass1,user2:pass2"
_COOKIE      = "skf_session"
_USER_COOKIE = "skf_user"

# Construir diccionario de usuarios
_USERS: dict[str, str] = {}
def _parse_users():
    global _USERS
    _USERS = {}
    if USERS_RAW:
        for pair in USERS_RAW.split(","):
            pair = pair.strip()
            if ":" in pair:
                u, p = pair.split(":", 1)
                _USERS[u.strip()] = p.strip()
    if not _USERS and APP_PASSWORD:
        _USERS["admin"] = APP_PASSWORD
_parse_users()


def _make_token(username: str) -> str:
    password = _USERS.get(username, "")
    return hmac.new(SECRET_KEY.encode(), f"{username}:{password}".encode(), hashlib.sha256).hexdigest()


def _valid_session(request: Request) -> tuple[bool, str]:
    """Devuelve (es_válida, username). Si no hay usuarios configurados, siempre True."""
    if not _USERS:
        return True, "anonymous"
    session_cookie = request.cookies.get(_COOKIE, "")
    user_hint = request.cookies.get(_USER_COOKIE, "")
    # Ruta rápida: comprobar usuario indicado en cookie
    if user_hint in _USERS:
        try:
            if hmac.compare_digest(session_cookie, _make_token(user_hint)):
                return True, user_hint
        except Exception:
            pass
    # Ruta lenta: probar todos los usuarios
    for username in _USERS:
        try:
            if hmac.compare_digest(session_cookie, _make_token(username)):
                return True, username
        except Exception:
            pass
    return False, ""


# Rate limiting para el login (en memoria)
_login_attempts: dict[str, list[float]] = {}
_RATE_LIMIT_MAX    = 5      # intentos fallidos máximos
_RATE_LIMIT_WINDOW = 900.0  # ventana de 15 minutos


def _get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _is_rate_limited(ip: str) -> bool:
    import time as _time
    now = _time.monotonic()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < _RATE_LIMIT_WINDOW]
    _login_attempts[ip] = attempts
    return len(attempts) >= _RATE_LIMIT_MAX


def _record_failed_login(ip: str):
    import time as _time
    _login_attempts.setdefault(ip, []).append(_time.monotonic())


def _clear_login_attempts(ip: str):
    _login_attempts.pop(ip, None)

# In-memory stores
_results_store: dict[str, dict] = {}
_progress_store: dict[str, dict] = {}
_claude_session_cache: dict[str, str] = {}  # clean_desc -> ref

# Global components
cache: CacheNormalizer | None = None
catalog: CatalogValidator | None = None
claude: ClaudeNormalizer | None = None
db: Database | None = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global cache, catalog, claude, db, _claude_session_cache

    model_path = DATA_DIR / "Modelo de aprendizaje.xlsx"
    catalog_path = DATA_DIR / "Data_ref.csv"

    try:
        cache = CacheNormalizer(model_path)
    except Exception as e:
        logger.error(f"No se pudo cargar el modelo de aprendizaje: {e}")

    try:
        catalog = CatalogValidator(catalog_path)
    except Exception as e:
        logger.error(f"No se pudo cargar el catálogo: {e}")

    if API_KEY and cache:
        examples = cache.get_examples(40)
        claude = ClaudeNormalizer(API_KEY, examples)
        logger.info("Claude API configurado correctamente")
    else:
        logger.warning("ANTHROPIC_API_KEY no configurada — sólo se usará la caché")

    # Inicializar base de datos (falla silenciosamente si no está configurada)
    if DB_HOST and DB_USER and DB_PASS and DB_NAME:
        try:
            db = init_db(DB_HOST, DB_USER, DB_PASS, DB_NAME, DB_PORT)
            if db.available():
                # Cargar caché persistente de Claude al arranque
                db_cache = db.get_all_claude_cache()
                _claude_session_cache.update(db_cache)
                logger.info(f"BD conectada · {len(db_cache)} entradas de caché Claude cargadas")
            else:
                logger.warning("BD configurada pero no disponible")
        except Exception as e:
            logger.warning(f"BD no disponible: {e}")
    else:
        logger.info("BD no configurada — funcionando sin persistencia MySQL")

    logger.info("Arranque completado")
    yield

    if db:
        db.close()


# ---------------------------------------------------------------------------
# Application
# ---------------------------------------------------------------------------
app = FastAPI(title="SKF Normalizador de Referencias", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # Rutas siempre públicas
        if path in ("/login", "/logout", "/api/status"):
            return await call_next(request)
        valid, _username = _valid_session(request)
        if not valid:
            if path.startswith("/api/"):
                return JSONResponse({"error": "No autorizado"}, status_code=401)
            return RedirectResponse(url="/login", status_code=302)
        return await call_next(request)


app.add_middleware(AuthMiddleware)


# ---------------------------------------------------------------------------
# Core processing logic
# ---------------------------------------------------------------------------

def _process_one(description: str, use_claude: bool = True, log_to_db: bool = False, username: str = "") -> dict:
    """Normaliza una descripción. Devuelve dict con ref, status, confidence, notes, source."""
    t0 = time.monotonic()

    if not description or not description.strip():
        return {"ref": None, "status": "NOT_FOUND", "confidence": "LOW", "notes": "Descripción vacía", "source": "none"}

    desc = description.strip()
    desc_clean = clean_description(desc)
    desc_norm = normalize_ref_candidate(desc_clean)

    result = None

    # --- 0. Lookup directo en catálogo ---
    if catalog:
        for candidate in ([desc_norm, desc_clean] if desc_norm != desc_clean else [desc_clean]):
            if not candidate:
                continue
            found_direct, matched_direct, score_direct = catalog.validate(candidate)
            if found_direct and score_direct == 100:
                result = {"ref": matched_direct, "status": "FOUND", "confidence": "HIGH",
                          "notes": "Referencia directa en catálogo", "source": "catalog"}
                break

    if result is None:
        # --- 1. Cache lookup ---
        if cache:
            ref_cache, score_cache, method_cache = cache.lookup(desc, cleaned=desc_clean)
        else:
            ref_cache, score_cache, method_cache = None, 0.0, "none"

        if method_cache == "exact" or (method_cache == "fuzzy" and score_cache >= 90):
            if catalog:
                found, matched, _cat_score = catalog.validate(ref_cache)
            else:
                found, matched = True, ref_cache

            confidence = "HIGH" if score_cache == 100 else "MEDIUM"
            status = "FOUND" if found else "REVIEW"
            notes = "Coincidencia exacta en modelo" if method_cache == "exact" else f"Coincidencia fuzzy {score_cache:.0f}% en modelo"
            if not found:
                notes += " (no validado en catálogo)"
            result = {"ref": matched or ref_cache, "status": status, "confidence": confidence, "notes": notes, "source": "cache"}

        # --- 2. Claude API ---
        elif use_claude and claude:
            ref_claude = claude.normalize_single(desc)

            if ref_claude and ref_claude.upper() not in ("UNKNOWN", "NAN", ""):
                if catalog:
                    found, matched_ref, cat_score = catalog.validate(ref_claude)
                else:
                    found, matched_ref, cat_score = True, ref_claude, 100.0

                if found:
                    confidence = "HIGH" if cat_score == 100 else "MEDIUM"
                    notes = "API Claude" if cat_score == 100 else f"API Claude + catálogo fuzzy {cat_score:.0f}%"
                    result = {"ref": matched_ref or ref_claude, "status": "FOUND", "confidence": confidence, "notes": notes, "source": "claude"}
                else:
                    if ref_cache and score_cache >= 75:
                        result = {
                            "ref": ref_cache,
                            "status": "REVIEW",
                            "confidence": "LOW",
                            "notes": f"Claude sugirió {ref_claude} (no en catálogo). Cache sugiere {ref_cache} ({score_cache:.0f}%)",
                            "source": "claude",
                        }
                    else:
                        result = {
                            "ref": ref_claude,
                            "status": "NOT_FOUND",
                            "confidence": "LOW",
                            "notes": f"Claude sugirió {ref_claude} pero no existe en catálogo",
                            "source": "claude",
                        }

        # --- 3. Fallback: weak cache suggestion ---
        if result is None:
            if ref_cache and score_cache >= 75:
                result = {"ref": ref_cache, "status": "REVIEW", "confidence": "LOW", "notes": f"Sólo coincidencia fuzzy débil {score_cache:.0f}%", "source": "cache"}
            else:
                result = {"ref": None, "status": "NOT_FOUND", "confidence": "LOW", "notes": "No identificado", "source": "none"}

    if log_to_db and db and db.available():
        ms = int((time.monotonic() - t0) * 1000)
        db.log_query(desc, desc_clean, result["ref"], result["status"],
                     result["confidence"], result["source"], result.get("notes", ""), ms, username)

    return result


def _read_file(content: bytes, filename: str) -> pd.DataFrame:
    if filename.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(content))
    detected = chardet.detect(content)
    encoding = detected.get("encoding") or "utf-8"
    try:
        return pd.read_csv(io.BytesIO(content), encoding=encoding, on_bad_lines="skip")
    except Exception:
        return pd.read_csv(io.BytesIO(content), encoding="latin-1", on_bad_lines="skip")


def _detect_desc_column(df: pd.DataFrame) -> str | None:
    keywords = ["descripcion", "descripción", "description", "referencia", "ref",
                "articulo", "artículo", "producto", "texto", "material", "item", "denominacion"]
    for col in df.columns:
        if any(k in col.lower() for k in keywords):
            return col
    for col in df.columns:
        if df[col].dtype == object:
            return col
    return None


def _build_output_excel(df_original: pd.DataFrame, results: list[dict]) -> bytes:
    df_out = df_original.copy()
    df_out["REF_NORMALIZADA"] = [r["ref"] if r["ref"] else "" for r in results]
    df_out["ESTADO"] = [r["status"] for r in results]
    df_out["CONFIANZA"] = [r["confidence"] for r in results]
    df_out["NOTAS"] = [r["notes"] for r in results]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Resultado")
        ws = writer.sheets["Resultado"]

        header_fill = PatternFill("solid", fgColor="0040A0")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

        status_col_idx = df_out.columns.get_loc("ESTADO") + 1
        fills = {
            "FOUND": PatternFill("solid", fgColor="C6EFCE"),
            "NOT_FOUND": PatternFill("solid", fgColor="FFC7CE"),
            "REVIEW": PatternFill("solid", fgColor="FFEB9C"),
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_val = row[status_col_idx - 1].value
            fill = fills.get(status_val)
            if fill:
                for cell in row:
                    cell.fill = fill

        for col_idx, col in enumerate(df_out.columns, 1):
            max_len = max(
                len(str(col)),
                df_out.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df_out) > 0 else 0,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    output.seek(0)
    return output.read()


def _build_output_csv(df_original: pd.DataFrame, results: list[dict]) -> bytes:
    df_out = df_original.copy()
    df_out["REF_NORMALIZADA"] = [r["ref"] if r["ref"] else "" for r in results]
    df_out["ESTADO"] = [r["status"] for r in results]
    df_out["CONFIANZA"] = [r["confidence"] for r in results]
    df_out["NOTAS"] = [r["notes"] for r in results]
    return df_out.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def _count_statuses(results: list[dict]) -> dict:
    counts = {"FOUND": 0, "NOT_FOUND": 0, "REVIEW": 0}
    for r in results:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    return counts


# ---------------------------------------------------------------------------
# PRO background processing
# ---------------------------------------------------------------------------

def _run_pro_sync(token: str):
    """Runs in thread pool. Updates _progress_store[token] as it goes."""
    global _claude_session_cache

    entry = _results_store.get(token)
    if not entry or not claude:
        _progress_store[token] = {"status": "error", "message": "Datos no disponibles"}
        return

    descriptions = entry["descriptions"]
    results = [dict(r) for r in entry["results"]]

    indices = [i for i, r in enumerate(results) if r["status"] in ("NOT_FOUND", "REVIEW")]
    total_pending = len(indices)

    _progress_store[token] = {"status": "processing", "current": 0, "total": total_pending, "improved": 0}

    if total_pending == 0:
        entry["results_pro"] = results
        entry["excel_pro"] = entry["excel_basic"]
        entry["csv_pro"] = entry["csv_basic"]
        counts = _count_statuses(results)
        _progress_store[token] = {
            "status": "done",
            "found": counts["FOUND"], "not_found": counts["NOT_FOUND"], "review": counts["REVIEW"],
            "improved": 0, "cost_eur": 0.0, "total": len(results),
        }
        return

    # Deduplicate: only send unique descriptions to Claude
    unique_descs = list(dict.fromkeys(descriptions[i] for i in indices))

    # Filter out session-cached results
    to_request = [d for d in unique_descs if clean_description(d) not in _claude_session_cache]

    def _progress_cb(current, total_chunks):
        pct_done = current / max(total_chunks, 1)
        _progress_store[token]["current"] = int(pct_done * total_pending)

    if to_request:
        claude_refs = claude.normalize_batch(to_request, progress_callback=_progress_cb)
        for desc, ref in zip(to_request, claude_refs):
            clean_key = clean_description(desc)
            _claude_session_cache[clean_key] = ref
            # Persist to DB cache
            if db and db.available():
                status_for_cache = "FOUND" if ref and ref.upper() not in ("UNKNOWN", "NAN", "") else "NOT_FOUND"
                db.save_claude_cache(clean_key, ref if status_for_cache == "FOUND" else None, status_for_cache)

    improved = 0
    for original_idx in indices:
        desc = descriptions[original_idx]
        ref_claude = _claude_session_cache.get(clean_description(desc), "UNKNOWN")

        if not ref_claude or ref_claude.upper() in ("UNKNOWN", "NAN", ""):
            continue
        if catalog:
            found, matched_ref, cat_score = catalog.validate(ref_claude)
        else:
            found, matched_ref, cat_score = True, ref_claude, 100.0

        if found:
            confidence = "HIGH" if cat_score == 100 else "MEDIUM"
            notes = "Claude API (Versión PRO)" if cat_score == 100 else f"Claude API + catálogo fuzzy {cat_score:.0f}%"
            results[original_idx] = {"ref": matched_ref or ref_claude, "status": "FOUND",
                                     "confidence": confidence, "notes": notes, "source": "claude"}
            improved += 1
        else:
            results[original_idx]["notes"] += f" | Claude sugirió: {ref_claude} (no en catálogo)"

    df = _read_file(entry["file_content"], entry["original_filename"])
    entry["results_pro"] = results
    entry["excel_pro"] = _build_output_excel(df, results)
    entry["csv_pro"] = _build_output_csv(df, results)

    counts = _count_statuses(results)
    cost_eur = claude.real_cost_eur() if claude else 0.0

    if db and db.available():
        est = _results_store.get(token, {}).get("_cost_estimated", 0.0)
        db.log_batch_pro(token, improved, est, cost_eur,
                         counts["FOUND"], counts["REVIEW"], counts["NOT_FOUND"])

    _progress_store[token] = {
        "status": "done",
        "found": counts["FOUND"], "not_found": counts["NOT_FOUND"], "review": counts["REVIEW"],
        "improved": improved, "cost_eur": cost_eur, "total": len(results),
    }
    logger.info(f"PRO batch done: token={token}, improved={improved}, cost=€{cost_eur}")
    if cost_eur >= COST_ALERT_EUR:
        logger.warning(f"⚠ ALERTA DE COSTE: gasto acumulado €{cost_eur:.4f} supera el umbral €{COST_ALERT_EUR}")


# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------

@app.get("/api/status")
async def status() -> JSONResponse:
    import subprocess, sys
    commit = ""
    try:
        commit = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"],
                                          cwd=str(BASE_DIR), stderr=subprocess.DEVNULL).decode().strip()
    except Exception:
        commit = "unknown"
    return JSONResponse({
        "cache_loaded": cache is not None,
        "catalog_loaded": catalog is not None,
        "claude_available": claude is not None,
        "cache_entries": len(cache.df) if cache else 0,
        "catalog_entries": len(catalog.refs) if catalog else 0,
        "db_available": db.available() if db else False,
        "session_cache_entries": len(_claude_session_cache),
        "active_jobs": len(_results_store),
        "user_count": len(_USERS) if _USERS else 0,
        "commit": commit,
        "uptime_jobs": len(_results_store),
    })


@app.post("/api/columns")
async def get_columns(file: UploadFile = File(...)) -> JSONResponse:
    content = await file.read()
    try:
        df = _read_file(content, file.filename)
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo: {e}")

    suggested = _detect_desc_column(df)

    # Send first 5 rows of suggested column for preview
    preview_rows = []
    if suggested and suggested in df.columns:
        preview_rows = df[suggested].astype(str).head(5).tolist()

    return JSONResponse({
        "columns": list(df.columns),
        "suggested": suggested,
        "rows": len(df),
        "preview_rows": preview_rows,
    })


@app.post("/api/normalize/single")
async def normalize_single(request: Request, description: str = Form(...)) -> JSONResponse:
    if not description.strip():
        raise HTTPException(400, "La descripción no puede estar vacía")
    _v, username = _valid_session(request)
    result = _process_one(description.strip(), log_to_db=True, username=username)
    return JSONResponse(result)


@app.post("/api/normalize/batch")
async def normalize_batch(
    file: UploadFile = File(...),
    col: Optional[str] = Form(None),
) -> JSONResponse:
    if not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(400, "El archivo debe ser Excel (.xlsx/.xls) o CSV (.csv)")

    content = await file.read()

    try:
        df = _read_file(content, file.filename)
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo: {e}")

    if df.empty:
        raise HTTPException(400, "El archivo está vacío")

    if len(df) > MAX_BATCH_ROWS:
        raise HTTPException(400, f"El archivo tiene {len(df)} filas. El límite es {MAX_BATCH_ROWS} filas por análisis. Divide el archivo en partes más pequeñas.")

    if col and col in df.columns:
        desc_col = col
    else:
        desc_col = _detect_desc_column(df)
        if desc_col is None:
            raise HTTPException(400, "No se pudo detectar la columna de descripciones")

    descriptions = df[desc_col].astype(str).tolist()

    # Basic processing with deduplication (cache + catalog only — no Claude)
    dedup_cache: dict[str, dict] = {}
    results = []
    for desc in descriptions:
        key = desc.strip().lower() if desc.strip() not in ("nan", "") else ""
        raw = desc.strip() if desc.strip() not in ("nan", "") else ""
        if key not in dedup_cache:
            dedup_cache[key] = _process_one(raw, use_claude=False)
        results.append(dedup_cache[key])

    excel_basic = _build_output_excel(df, results)
    csv_basic = _build_output_csv(df, results)

    token = str(uuid.uuid4())
    orig_name = file.filename.rsplit(".", 1)[0]
    _results_store[token] = {
        "descriptions": descriptions,
        "results": results,
        "results_pro": None,
        "excel_basic": excel_basic,
        "csv_basic": csv_basic,
        "excel_pro": None,
        "csv_pro": None,
        "file_content": content,
        "original_filename": file.filename,
        "desc_col": desc_col,
        "orig_name": orig_name,
    }

    counts = _count_statuses(results)

    if db and db.available():
        db.log_batch(token, file.filename, len(results),
                     counts["FOUND"], counts["REVIEW"], counts["NOT_FOUND"])

    return JSONResponse({
        "token": token,
        "total": len(results),
        "found": counts["FOUND"],
        "not_found": counts["NOT_FOUND"],
        "review": counts["REVIEW"],
        "desc_col": desc_col,
        "claude_available": claude is not None,
    })


@app.get("/api/estimate/{token}")
async def estimate_cost(token: str) -> JSONResponse:
    entry = _results_store.get(token)
    if not entry:
        raise HTTPException(404, "Token no encontrado")

    pending = [r for r in entry["results"] if r["status"] in ("NOT_FOUND", "REVIEW")]

    # Deduplicate estimate (unique descriptions only)
    unique_pending = set(
        entry["descriptions"][i].strip().lower()
        for i, r in enumerate(entry["results"])
        if r["status"] in ("NOT_FOUND", "REVIEW")
    )
    # Subtract already session-cached
    uncached = {d for d in unique_pending if clean_description(d) not in _claude_session_cache}
    n_api_calls = len(uncached)

    batches = max(1, -(-n_api_calls // 10))
    input_tokens = batches * 2700
    output_tokens = batches * 150
    cost_usd = (input_tokens * 3 + output_tokens * 15) / 1_000_000
    cost_usd_max = cost_usd * 1.35
    cost_eur = cost_usd * 0.92
    cost_eur_max = cost_usd_max * 0.92

    return JSONResponse({
        "pending": len(pending),
        "unique_api_calls": n_api_calls,
        "session_cached": len(unique_pending) - n_api_calls,
        "claude_available": claude is not None,
        "cost_usd": round(cost_usd, 4),
        "cost_eur": round(cost_eur, 4),
        "cost_usd_max": round(cost_usd_max, 4),
        "cost_eur_max": round(cost_eur_max, 4),
    })


@app.post("/api/normalize/batch/pro/{token}")
async def normalize_batch_pro_start(token: str) -> JSONResponse:
    """Inicia el procesamiento PRO en background. Responde inmediatamente."""
    entry = _results_store.get(token)
    if not entry:
        raise HTTPException(404, "Token no encontrado")
    if not claude:
        raise HTTPException(400, "Claude API no configurada. Añade ANTHROPIC_API_KEY al archivo .env")

    # Avoid double-start
    existing = _progress_store.get(token, {})
    if existing.get("status") == "processing":
        return JSONResponse({"status": "already_processing"})

    _progress_store[token] = {"status": "processing", "current": 0, "total": 0, "improved": 0}
    asyncio.create_task(asyncio.to_thread(_run_pro_sync, token))
    return JSONResponse({"status": "started"})


@app.get("/api/progress/{token}")
async def get_progress(token: str) -> JSONResponse:
    progress = _progress_store.get(token)
    if not progress:
        raise HTTPException(404, "No hay progreso registrado")
    return JSONResponse(progress)


@app.get("/api/download/{token}")
async def download_result(
    token: str,
    version: str = "basic",
    fmt: str = "xlsx",
    status_filter: str = "all",
) -> StreamingResponse:
    entry = _results_store.get(token)
    if not entry:
        raise HTTPException(404, "Resultado no encontrado o expirado")

    if version == "pro":
        results = entry.get("results_pro")
        if not results:
            raise HTTPException(400, "Versión PRO no disponible todavía")
        orig_data = entry["excel_pro"]
        orig_csv = entry.get("csv_pro")
        suffix = "PRO"
    else:
        results = entry["results"]
        orig_data = entry["excel_basic"]
        orig_csv = entry.get("csv_basic")
        suffix = "basico"

    orig_name = entry["orig_name"]

    # Apply status filter if requested
    if status_filter != "all" and results:
        df = _read_file(entry["file_content"], entry["original_filename"])
        filtered_results = []
        filtered_rows = []
        for i, r in enumerate(results):
            if r["status"] == status_filter.upper():
                filtered_results.append(r)
                filtered_rows.append(i)
        df_filtered = df.iloc[filtered_rows].reset_index(drop=True)
        if fmt == "csv":
            data = _build_output_csv(df_filtered, filtered_results)
        else:
            data = _build_output_excel(df_filtered, filtered_results)
        suffix = f"{suffix}_{status_filter}"
    else:
        data = orig_csv if fmt == "csv" else orig_data

    if fmt == "csv":
        filename = f"resultado_{orig_name}_{suffix}.csv"
        media_type = "text/csv"
    else:
        filename = f"resultado_{orig_name}_{suffix}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        io.BytesIO(data),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Stats, feedback y utilidades
# ---------------------------------------------------------------------------

@app.get("/api/stats")
async def get_stats() -> JSONResponse:
    """Dashboard de estadísticas de uso."""
    if db and db.available():
        stats = db.get_detailed_stats()
    else:
        stats = {
            "total_queries": 0, "found": 0, "not_found": 0, "review": 0,
            "from_cache": 0, "from_claude": 0, "from_catalog": 0,
            "avg_response_ms": 0, "claude_cache_entries": 0,
            "top_refs": [], "queries_by_day": [],
            "total_batches": 0, "total_batch_rows": 0, "pro_batches": 0,
            "total_cost_eur": 0.0,
            "feedback_total": 0, "feedback_positive": 0, "feedback_negative": 0,
        }
    # Enriquecer con datos en memoria
    stats["session_cache_entries"] = len(_claude_session_cache)
    stats["active_jobs"] = len(_results_store)
    # Número de usuarios configurados
    stats["user_count"] = len(_USERS) if _USERS else 0
    return JSONResponse(stats)


@app.post("/api/feedback")
async def post_feedback(
    description: str = Form(...),
    ref: str = Form(""),
    thumbs_up: str = Form(...),   # "true" / "false"
    source: str = Form(""),
) -> JSONResponse:
    """Registra el feedback del usuario (👍/👎) sobre un resultado."""
    positive = thumbs_up.lower() in ("true", "1", "yes")
    if db and db.available():
        db.log_feedback(description, ref or None, positive, source)
    return JSONResponse({"ok": True})


@app.get("/api/my-ip")
async def get_my_ip() -> JSONResponse:
    """Devuelve la IP saliente del servidor (útil para whitelistear en MySQL)."""
    def _fetch():
        try:
            with urllib.request.urlopen("https://api.ipify.org", timeout=5) as r:
                return r.read().decode().strip()
        except Exception as e:
            return f"error: {e}"
    ip = await asyncio.to_thread(_fetch)
    return JSONResponse({"outbound_ip": ip})


@app.post("/api/cache/clear")
async def clear_cache(request: Request) -> JSONResponse:
    """Vacía la caché de sesión de Claude sin reiniciar el servidor."""
    global _claude_session_cache
    _valid, username = _valid_session(request)
    n = len(_claude_session_cache)
    _claude_session_cache.clear()
    logger.info(f"Caché de sesión vaciada por {username}: {n} entradas eliminadas")
    return JSONResponse({"cleared": n, "ok": True})


@app.post("/api/cache/correct")
async def correct_cache(
    request: Request,
    description: str = Form(...),
    ref_correct: str = Form(...),
) -> JSONResponse:
    """Guarda una corrección manual en la caché de Claude."""
    global _claude_session_cache
    _valid, username = _valid_session(request)
    desc_clean = clean_description(description.strip())
    ref_clean = ref_correct.strip().upper()
    _claude_session_cache[desc_clean] = ref_clean
    if db and db.available():
        db.save_claude_cache(desc_clean, ref_clean, "FOUND")
        logger.info(f"Corrección manual guardada por {username}: '{desc_clean}' → {ref_clean}")
    return JSONResponse({"ok": True, "desc_clean": desc_clean, "ref": ref_clean})


@app.get("/api/cache/export")
async def export_cache() -> StreamingResponse:
    """Exporta toda la caché de correcciones como Excel para merge con el modelo base."""
    rows = []
    if db and db.available():
        rows = db.get_full_cache_for_export()
    if not rows:
        rows = [{"description_clean": k, "ref_result": v, "status": "FOUND", "used_count": 1, "last_used_at": None}
                for k, v in _claude_session_cache.items()]
    df_export = pd.DataFrame([{
        "DESCRIPCION_LIMPIA": r["description_clean"],
        "REF_NORMALIZADA": r["ref_result"],
        "USOS": r.get("used_count", 1),
        "ULTIMA_VEZ": str(r["last_used_at"]) if r.get("last_used_at") else "",
    } for r in rows])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Correcciones")
        ws = writer.sheets["Correcciones"]
        hfill = PatternFill("solid", fgColor="0040A0")
        hfont = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(col[0].value or "")),
                    max((len(str(c.value or "")) for c in col[1:]), default=0)) + 3, 60)
    output.seek(0)
    from datetime import date
    filename = f"correcciones_modelo_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/normalize/batch/rerun/{token}")
async def normalize_batch_rerun(token: str, status_filter: str = Form("REVIEW")) -> JSONResponse:
    """Re-procesa solo los items con un estado específico (REVIEW por defecto) usando la caché actualizada."""
    entry = _results_store.get(token)
    if not entry:
        raise HTTPException(404, "Token no encontrado")

    results = list(entry.get("results_pro") or entry["results"])
    descriptions = entry["descriptions"]

    rerun_indices = [i for i, r in enumerate(results) if r["status"] == status_filter.upper()]
    improved = 0
    for idx in rerun_indices:
        desc = descriptions[idx]
        desc_clean = clean_description(desc)
        # Check session cache first
        if desc_clean in _claude_session_cache:
            ref = _claude_session_cache[desc_clean]
            if ref and ref.upper() not in ("UNKNOWN", "NAN", ""):
                if catalog:
                    found, matched_ref, cat_score = catalog.validate(ref)
                else:
                    found, matched_ref, cat_score = True, ref, 100.0
                if found:
                    results[idx] = {"ref": matched_ref or ref, "status": "FOUND",
                                    "confidence": "HIGH", "notes": "Corrección manual desde caché", "source": "cache"}
                    improved += 1
        else:
            # Re-run through full pipeline with claude
            new_result = _process_one(desc, use_claude=True, log_to_db=False)
            if new_result["status"] == "FOUND":
                results[idx] = new_result
                improved += 1

    df = _read_file(entry["file_content"], entry["original_filename"])
    if entry.get("results_pro"):
        entry["results_pro"] = results
        entry["excel_pro"] = _build_output_excel(df, results)
        entry["csv_pro"] = _build_output_csv(df, results)
    else:
        entry["results"] = results
        entry["excel_basic"] = _build_output_excel(df, results)
        entry["csv_basic"] = _build_output_csv(df, results)

    counts = _count_statuses(results)
    return JSONResponse({
        "improved": improved,
        "found": counts["FOUND"],
        "not_found": counts["NOT_FOUND"],
        "review": counts["REVIEW"],
        "total": len(results),
    })


# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------

_LOGIN_HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SKF · Acceso</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    background: #001f5b;
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
  }
  .card {
    background: #fff;
    border-radius: 12px;
    padding: 2.5rem 2rem;
    width: 100%;
    max-width: 360px;
    box-shadow: 0 8px 32px rgba(0,0,0,0.3);
    text-align: center;
  }
  .logo {
    font-size: 1.6rem;
    font-weight: 800;
    color: #003087;
    letter-spacing: .08em;
    margin-bottom: .25rem;
  }
  .subtitle {
    font-size: .85rem;
    color: #666;
    margin-bottom: 2rem;
  }
  label {
    display: block;
    text-align: left;
    font-size: .8rem;
    font-weight: 600;
    color: #444;
    margin-bottom: .4rem;
  }
  input[type=password], input[type=text] {
    width: 100%;
    padding: .75rem 1rem;
    border: 1.5px solid #ddd;
    border-radius: 8px;
    font-size: .95rem;
    outline: none;
    transition: border-color .2s;
    margin-bottom: 1.25rem;
  }
  input[type=password]:focus, input[type=text]:focus { border-color: #003087; }
  button {
    width: 100%;
    padding: .8rem;
    background: #003087;
    color: #fff;
    border: none;
    border-radius: 8px;
    font-size: 1rem;
    font-weight: 600;
    cursor: pointer;
    transition: background .2s;
  }
  button:hover { background: #0040b0; }
  .error {
    background: #fff0f0;
    color: #c00;
    border-radius: 6px;
    padding: .6rem .9rem;
    font-size: .85rem;
    margin-bottom: 1rem;
  }
</style>
</head>
<body>
<div class="card">
  <div class="logo">SKF</div>
  <div class="subtitle">Normalizador de Referencias</div>
  {error_block}
  <form method="post" action="/login">
    {user_field}
    <label for="pwd">Contraseña de acceso</label>
    <input type="password" id="pwd" name="password" autofocus placeholder="••••••••">
    <button type="submit">Entrar</button>
  </form>
</div>
</body>
</html>"""


def _render_login(error: str = "", show_user_field: bool = False) -> str:
    user_field = """
    <label for="usr">Usuario</label>
    <input type="text" id="usr" name="username" autocomplete="username" placeholder="usuario">
    """ if show_user_field else '<input type="hidden" name="username" value="">'
    html = _LOGIN_HTML.replace("{error_block}", error).replace("{user_field}", user_field)
    return html


@app.get("/login", response_class=HTMLResponse, include_in_schema=False)
async def login_page():
    show_users = len(_USERS) > 1
    return HTMLResponse(_render_login(show_user_field=show_users))


@app.post("/login", response_class=HTMLResponse, include_in_schema=False)
async def login_submit(request: Request, password: str = Form(...), username: str = Form("")):
    ip = _get_client_ip(request)
    show_users = len(_USERS) > 1

    # Comprobar rate limit
    if _is_rate_limited(ip):
        err = '<div class="error">Demasiados intentos fallidos. Espera 15 minutos e inténtalo de nuevo.</div>'
        return HTMLResponse(_render_login(err, show_users), status_code=429)

    # Verificar credenciales
    matched_user: str | None = None
    if _USERS:
        for uname, upass in _USERS.items():
            # Si se indicó usuario, sólo comprobamos ese; si no, probamos todos
            if username and username != uname:
                continue
            try:
                if hmac.compare_digest(password, upass):
                    matched_user = uname
                    break
            except Exception:
                pass

    if matched_user is None:
        _record_failed_login(ip)
        err = '<div class="error">Credenciales incorrectas.</div>'
        return HTMLResponse(_render_login(err, show_users), status_code=401)

    _clear_login_attempts(ip)
    response = RedirectResponse(url="/", status_code=302)
    response.set_cookie(
        key=_COOKIE, value=_make_token(matched_user),
        httponly=True, samesite="lax", max_age=60 * 60 * 24 * 30,
    )
    # Cookie de usuario (no sensible, sólo para display)
    response.set_cookie(
        key=_USER_COOKIE, value=matched_user,
        httponly=False, samesite="lax", max_age=60 * 60 * 24 * 30,
    )
    return response


@app.get("/logout", include_in_schema=False)
async def logout():
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie(_COOKIE)
    response.delete_cookie(_USER_COOKIE)
    return response


# ---------------------------------------------------------------------------
# Serve frontend (must be last)
# ---------------------------------------------------------------------------
app.mount("/", StaticFiles(directory=str(BASE_DIR / "static"), html=True), name="static")


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=True)
